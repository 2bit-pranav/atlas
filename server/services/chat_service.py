import asyncio
from pathlib import Path
from typing import Any, AsyncGenerator, Dict, List, Union, Optional

from autogen_agentchat.base import TaskResult
from autogen_agentchat.messages import (
    ModelClientStreamingChunkEvent,
    MultiModalMessage,
    TextMessage,
    ThoughtEvent,
)
from autogen_core import Image

from agent.agent import create_atlas_agent
from agent.config import get_cloud_model, get_local_model
from ..managers import chat_session_manager, skill_manager


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}
TEXT_EXTENSIONS = {".txt", ".md", ".py", ".json", ".csv", ".log", ".html", ".xml", ".yml", ".yaml", ".js", ".ts"}
PDF_EXTENSIONS = {".pdf"}
DOCX_EXTENSIONS = {".docx"}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}

_THOUGHT_PREFIX = "<|atlas_thought|>"


class ChatService:
    """Handles file operations, prompt construction, and chat orchestration."""

    @staticmethod
    def validate_path(path: str) -> Optional[Path]:
        if not path or not path.strip():
            return None
        candidate = path.strip().strip('"')
        p = Path(candidate).expanduser().resolve(strict=False)
        return p if p.exists() and p.is_file() else None

    @classmethod
    def extract_file_content(cls, path: Path, display_name: Optional[str] = None) -> Union[str, Image, None]:
        name = display_name or path.name
        suffix = path.suffix.lower()
        abs_path = str(path.resolve())

        if suffix in IMAGE_EXTENSIONS:
            try:
                return Image.from_file(abs_path)
            except Exception as e:
                return f"[Error loading image {name}: {e}]"

        if suffix in TEXT_EXTENSIONS:
            try:
                content = path.read_text(encoding="utf-8", errors="replace")
                return f"[Attached Text File: {name}]\nFILE_PATH_ON_DISK: {abs_path}\n--- CONTENT ---\n{content}"
            except Exception as e:
                return f"[Error reading text file {name}: {e}]"

        if suffix in PDF_EXTENSIONS:
            try:
                from pypdf import PdfReader
                reader = PdfReader(abs_path)
                pages = [page.extract_text() or "" for page in reader.pages]
                return f"[Attached PDF Document: {name}]\nFILE_PATH_ON_DISK: {abs_path}\n--- EXTRACTED TEXT ---\n" + "\n".join(pages).strip()
            except Exception as e:
                return f"[Error reading PDF {name}: {e}]"

        if suffix in DOCX_EXTENSIONS:
            try:
                import docx
                doc = docx.Document(abs_path)
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                return f"[Attached Word Document: {name}]\nFILE_PATH_ON_DISK: {abs_path}\n--- EXTRACTED TEXT ---\n" + "\n".join(paragraphs)
            except Exception as e:
                return f"[Error reading DOCX {name}: {e}]"

        if suffix in EXCEL_EXTENSIONS:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(abs_path, data_only=True)
                sheets = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = ["\t".join(str(c) if c is not None else "" for c in r) for r in ws.iter_rows(values_only=True) if any(r)]
                    if rows:
                        sheets.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
                return f"[Attached Excel Spreadsheet: {name}]\nFILE_PATH_ON_DISK: {abs_path}\n" + "\n\n".join(sheets)
            except Exception as e:
                return f"[Error reading Excel {name}: {e}]"

        return f"[Unsupported file type {suffix} for file {name}]"

    @classmethod
    def build_task_payload(
        cls,
        prompt: str,
        attachments: Optional[List[Union[str, Dict[str, str]]]],
        chat_id: str,
    ) -> Union[str, MultiModalMessage]:

        content_items: List[Union[str, Image]] = []
        clean_prompt = prompt.strip()

        # Resolve all attachment paths upfront so we can pass them into skill context
        resolved_attachment_paths: List[str] = []
        if attachments:
            for item in attachments:
                raw_path = item.get("path", "") if isinstance(item, dict) else str(item)
                p = cls.validate_path(raw_path)
                if p:
                    resolved_attachment_paths.append(str(p.resolve()))

        skill_context = skill_manager.context_for_prompt(
            clean_prompt,
            uploaded_file_paths=resolved_attachment_paths or None,
        )
        if skill_context:
            content_items.append(
                "Available or requested skill context:\n" + skill_context
            )

        if clean_prompt:
            content_items.append(clean_prompt)

        if attachments:
            for item in attachments:
                if isinstance(item, dict):
                    raw_path = item.get("path", "")
                    display_name = item.get("name")
                else:
                    raw_path = str(item)
                    display_name = None

                p = cls.validate_path(raw_path)
                if not p:
                    continue

                fname = display_name or p.name
                extracted = cls.extract_file_content(p, display_name=fname)
                if isinstance(extracted, Image):
                    content_items.append(f"[Attached Image: {fname}]")
                    content_items.append(extracted)
                elif isinstance(extracted, str):
                    content_items.append(extracted)

        if not content_items:
            raise ValueError("Message must include text or a valid attachment.")

        if len(content_items) == 1 and isinstance(content_items[0], str):
            return content_items[0]

        return MultiModalMessage(content=content_items, source="user")

    @classmethod
    async def process_chat(
        cls,
        prompt: str,
        chat_id: Optional[str] = None,
        thinking_budget: int = 0,
        use_cloud: bool = False,
        attachments: Optional[List[Union[str, Dict[str, str]]]] = None,
    ) -> AsyncGenerator[Dict[str, Any], None]:

        # Initialize or retrieve session
        active_chat = chat_session_manager.create_session(chat_id)
        session = chat_session_manager.get_session(active_chat)

        clean_prompt = prompt.strip()

        # Set chat title from first message
        if session.get("title") == "New Chat" or not session.get("messages"):
            if clean_prompt:
                first_line = clean_prompt.split("\n")[0].strip()
                title = first_line[:35] + ("..." if len(first_line) > 35 else "")
            elif attachments:
                first_att = attachments[0]
                fname = first_att.get("name") if isinstance(first_att, dict) else Path(str(first_att)).name
                title = f"File: {fname}"
            else:
                title = "Chat Session"
            chat_session_manager.set_title(active_chat, title)

        # Build attachment metadata for user message
        user_attachment_meta = []
        if attachments:
            for a in attachments:
                fname = a.get("name") if isinstance(a, dict) else Path(str(a)).name
                is_img = any(fname.lower().endswith(ext) for ext in IMAGE_EXTENSIONS)
                user_attachment_meta.append({
                    "name": fname,
                    "type": "image" if is_img else "document",
                })

        # Record user message
        chat_session_manager.add_message(
            active_chat,
            role="user",
            content=clean_prompt,
            attachments=user_attachment_meta,
        )

        from agent.browser_agent.tools import set_browser_chat_id, set_terminal_callback

        # Sentinel value to signal stream completion
        _STREAM_DONE = object()

        terminal_queue: asyncio.Queue[Any] = asyncio.Queue()
        msg_queue: asyncio.Queue = asyncio.Queue()
        loop = asyncio.get_running_loop()

        def terminal_cb(msg: str):
            loop.call_soon_threadsafe(terminal_queue.put_nowait, msg)

        set_terminal_callback(terminal_cb)
        set_browser_chat_id(active_chat)

        model_client = get_cloud_model() if use_cloud else get_local_model(thinking_budget=thinking_budget)
        atlas = create_atlas_agent(model_client=model_client)

        # Load previous agent state if available
        agent_state = chat_session_manager.get_agent_state(active_chat)
        if agent_state:
            await atlas.load_state(agent_state)

        yield {"type": "meta", "chat_id": active_chat}
        yield {"type": "status", "status": "running", "label": "Preparing prompt"}

        task = cls.build_task_payload(prompt, attachments, active_chat)
        yield {"type": "status", "status": "running", "label": "Running Atlas"}

        accumulated_content = ""
        accumulated_thought = ""
        chunk_yielded = False

        # Run atlas.run_stream() as a background task feeding msg_queue.
        # This frees the generator to poll BOTH queues concurrently, so terminal
        # step logs from the browser worker thread stream through in real-time
        # instead of being batched until the browser tool returns.
        async def _stream_to_queue():
            try:
                async for message in atlas.run_stream(task=task):
                    await msg_queue.put(message)
            finally:
                await msg_queue.put(_STREAM_DONE)

        stream_task = asyncio.create_task(_stream_to_queue())

        try:
            while True:
                # Drain any pending terminal logs first (non-blocking)
                while not terminal_queue.empty():
                    log_item = terminal_queue.get_nowait()
                    if isinstance(log_item, dict):
                        yield log_item
                        if log_item.get("type") == "browser_handoff":
                            yield {
                                "type": "terminal",
                                "content": f"USER HANDOFF REQUIRED: {log_item.get('question', 'Browser input required.')}",
                            }
                    else:
                        yield {"type": "terminal", "content": str(log_item)}

                # Wait for next atlas message with a short timeout so we can
                # re-check terminal_queue while the browser tool is running.
                try:
                    message = await asyncio.wait_for(msg_queue.get(), timeout=0.1)
                except asyncio.TimeoutError:
                    # No atlas message yet; loop back to drain terminal logs
                    continue

                if message is _STREAM_DONE:
                    break

                if isinstance(message, ModelClientStreamingChunkEvent):
                    content = message.content
                    if not content:
                        continue
                    chunk_yielded = True
                    if content.startswith(_THOUGHT_PREFIX):
                        thought_text = content[len(_THOUGHT_PREFIX):]
                        if thought_text:
                            accumulated_thought += thought_text
                            yield {"type": "thought", "content": thought_text}
                    else:
                        accumulated_content += content
                        yield {"type": "chunk", "content": content}

                elif isinstance(message, ThoughtEvent):
                    if message.content:
                        accumulated_thought += message.content
                        yield {"type": "thought", "content": message.content}

                elif isinstance(message, TaskResult) and not chunk_yielded:
                    for msg in reversed(message.messages):
                        if isinstance(msg, TextMessage) and msg.source != "user":
                            accumulated_content += msg.content
                            yield {"type": "chunk", "content": msg.content}
                            break

            # Drain any remaining terminal logs after stream ends
            while not terminal_queue.empty():
                log_item = terminal_queue.get_nowait()
                yield log_item if isinstance(log_item, dict) else {"type": "terminal", "content": str(log_item)}

        finally:
            set_terminal_callback(None)
            set_browser_chat_id(None)
            if not stream_task.done():
                stream_task.cancel()
                try:
                    await stream_task
                except (asyncio.CancelledError, Exception):
                    pass

            yield {"type": "status", "status": "completed", "label": "Completed"}

        # Record assistant message
        chat_session_manager.add_message(
            active_chat,
            role="assistant",
            content=accumulated_content,
            thought=accumulated_thought if accumulated_thought else None,
        )

        # Save agent state for continued conversation
        agent_state = await atlas.save_state()
        chat_session_manager.set_agent_state(active_chat, agent_state)